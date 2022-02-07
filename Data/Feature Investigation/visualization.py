from pandas import read_excel, DataFrame
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from statsmodels.tsa.seasonal import seasonal_decompose

def seperateYears(dates, flowRates):
    '''Returns a list of numpy arrays, contained in which are the flow rates seperated by year'''
    year2017 = []
    year2018 = []
    year2019 = []
    year2020 = []
    year2021 = []
    year2022 = []

    for i in range(len(dates)):
        if dates[i][:4] == "2017":
            year2017.append(flowRates[i])
        elif dates[i][:4] == "2018":
            year2018.append(flowRates[i])
        elif dates[i][:4] == "2019":
            year2019.append(flowRates[i])
        elif dates[i][:4] == "2020":
            year2020.append(flowRates[i])
        elif dates[i][:4] == "2021":
            year2021.append(flowRates[i])
        else:
            year2022.append(flowRates[i])
        

    year2017 = np.array(year2017)
    year2018 = np.array(year2018)
    year2019 = np.array(year2019)
    year2020 = np.array(year2020)
    year2021 = np.array(year2021)
    year2022 = np.array(year2022)
    # print(np.size(year2017), np.size(year2018), np.size(year2019), np.size(year2020), np.size(year2021), np.size(year2022))
    return [year2017, year2018, year2019, year2020, year2021, year2022]

def removeBadData(dates, flowRates):
    '''Removes data where null values or a total influent rate < 4 is present.
    Returns: dates, flowRates as np arrays'''
    badIncidies = []
    for i in range(len(flowRates)):
        if flowRates[i] < 4:
            badIncidies.append(i)
    return np.delete(dates, badIncidies), np.delete(flowRates, badIncidies)        


def createTotalTimeSeriesGraph(dates, flowRates):
    '''Graphs the total influent rate for all the data per hour, adding vertical lines to
    indicate year seperation'''
    years = seperateYears(dates, flowRates)
    plt.plot(list(range(0, len(dates))), flowRates, linewidth=0.25)
    # plt.plot(list(range(0, len(dates))), len(dates)*[5], linewidth=0.5)
    plt.plot(300*[len(years[0])], np.linspace(0, 115, 300), "m--", linewidth=2)
    plt.plot(300*[len(years[0]) + len(years[1])], np.linspace(0, 115, 300), "m--", linewidth=2)
    plt.plot(300*[len(years[0]) + len(years[1]) + len(years[2])], np.linspace(0, 115, 300), "m--", linewidth=2)
    plt.plot(300*[len(years[0]) + len(years[1]) + len(years[2]) + len(years[3])], np.linspace(0, 115, 300), "m--", linewidth=2)
    plt.plot(300*[len(years[0]) + len(years[1]) + len(years[2]) + len(years[3]) + len(years[4])], np.linspace(0, 115, 300), "m--", linewidth=2)
    plt.show()

def createYearlyTimeSeriesGraph(dates, flowRates):
    '''Graphs each year seperately on a single graph with legend.
    Note that year that is a leap year has extra data than others, and
    2022 has much less data than others'''
    years = seperateYears(dates, flowRates)
    plt.plot(np.linspace(0, len(years[0]), len(years[0])), years[0], label = "2017", linewidth=0.35)
    # # plt.plot(np.linspace(0, len(years[1]), len(years[1])), years[1], label = "2018", linewidth=0.35)
    plt.plot(np.linspace(0, len(years[2]), len(years[2])), years[2], label = "2019", linewidth=0.35)
    # plt.plot(np.linspace(0, len(years[3]), len(years[3])), years[3], label = "2020", linewidth=0.35)
    # plt.plot(np.linspace(0, len(years[4]), len(years[4])), years[4], label = "2021", linewidth=0.35)
    # # plt.plot(np.linspace(0, len(years[5]), len(years[5])), years[5], label = "2022", linewidth=0.35)
    plt.legend()
    plt.show()


def decomposeTimeSeriesSeasonal(dates, flowRates):
    dates, flowRates = removeBadData(dates, flowRates)
    # combined = np.vstack((dates, flowRates)).T
    # df = DataFrame(combined, columns=["DateTime", "SWTP Total Influent Rate"])

    decomp = seasonal_decompose(flowRates, model='multiplicative', period=365*24)
    # decomp.plot()
    # plt.show()
    
    plt.plot(np.linspace(0, len(dates), len(dates)), decomp.seasonal)
    top = 2.3
    bottom = .45
    years = years = seperateYears(dates, flowRates)
    plt.plot(300*[len(years[0])], np.linspace(bottom, top, 300), "m--", linewidth=2)
    plt.plot(300*[len(years[0]) + len(years[1])], np.linspace(bottom, top, 300), "m--", linewidth=2)
    plt.plot(300*[len(years[0]) + len(years[1]) + len(years[2])], np.linspace(bottom, top, 300), "m--", linewidth=2)
    plt.plot(300*[len(years[0]) + len(years[1]) + len(years[2]) + len(years[3])], np.linspace(bottom, top, 300), "m--", linewidth=2)
    plt.plot(300*[len(years[0]) + len(years[1]) + len(years[2]) + len(years[3]) + len(years[4])], np.linspace(bottom, top, 300), "m--", linewidth=2)
    plt.show()


def colorCells():
    '''Goes through each sheet in the file 'SWTP Total Data.xlsx', 
    highlighting rows where Null values are present or
    the total influent flow is less than 4'''
    wb = load_workbook(filename="SWTP Total Data.xlsx", data_only=True)
    yellowFill = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type='solid')

    for i in range(1, len(wb.sheetnames)):
        ws = wb[wb.sheetnames[i]]
        for i in range(5, 9000):
            if ws["E" + str(i)].value == "Null" or ws["H" + str(i)].value == "Null" or ws["K" + str(i)].value == "Null" or ws["N" + str(i)].value == "Null":
                for rows in ws.iter_rows(min_row = i, max_row = i, min_col = 1, max_col = 17):
                    for cell in rows:
                        cell.fill = yellowFill
            else:
                try:           
                    qthValue = float(ws["E" + str(i)].value) + float(ws["H" + str(i)].value) + float(ws["K" + str(i)].value) + float(ws["N" + str(i)].value)
                    if qthValue < 4:
                        for rows in ws.iter_rows(min_row = i, max_row = i, min_col = 1, max_col = 17):
                            for cell in rows:
                                cell.fill = yellowFill
                except TypeError:
                    pass

    ws = wb["Total"]
    for i in range(2, 44044):
        if float(ws["B" + str(i)].value) < 4:
            for rows in ws.iter_rows(min_row = i, max_row = i, min_col = 1, max_col = 2):
                for cell in rows:
                    cell.fill = yellowFill
            # print(i)

    wb.save("SWTP Total Data.xlsx")


def main():
    df = read_excel("SWTP Total Data - Copy.xlsx", sheet_name="Total", engine="openpyxl") #openpyxl is a library itself, helps to read xlsx files
    dates = df["DateTime"].copy(deep=True)
    flowRates = df["SWTP Total Influent Flow"].copy(deep=True)

    dates = np.array(list(map(str, dates)))
    flowRates = np.array(flowRates)


    # createTotalTimeSeriesGraph(dates, flowRates)
    # createYearlyTimeSeriesGraph(dates, flowRates)
    # decomposeTimeSeriesSeasonal(dates, flowRates)

    # colorCells()

    # flowRates.sort()
    # count = 0
    # index = 0
    # while count < 10:
    #     if flowRates[index] > 0.1:
    #         count += 1
    #         print(flowRates[index])
    #     index += 1



    
        
if __name__ == "__main__":
    main()